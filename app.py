from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from flask_sqlalchemy import SQLAlchemy
import os
import json
import pandas as pd
from datetime import datetime
import io
from functools import wraps

app = Flask(__name__)

# Configuración básica
app.config['SECRET_KEY'] = 'tu_clave_secreta_aqui_muy_segura_2025'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///solicitudes.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Inicializar SQLAlchemy
db = SQLAlchemy(app)

# Modelo de la base de datos
class Solicitud(db.Model):
    __tablename__ = 'solicitudes'
    
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    nombres_completos = db.Column(db.String(200), nullable=False)
    dni = db.Column(db.String(8), unique=True, nullable=False)
    profesion = db.Column(db.String(100), nullable=False)
    empresa = db.Column(db.String(200), nullable=False)
    cargo = db.Column(db.String(100), nullable=False)
    sector = db.Column(db.String(100), nullable=False)
    fecha_nacimiento = db.Column(db.Date, nullable=False)
    estado = db.Column(db.String(20), default='Pendiente', nullable=False)
    evaluador = db.Column(db.String(100), nullable=True)
    observaciones = db.Column(db.Text, nullable=True)
    fecha_evaluacion = db.Column(db.DateTime, nullable=True)
    
    def to_dict(self):
        """Convertir el objeto a diccionario"""
        return {
            'id': self.id,
            'fecha': self.fecha.strftime('%Y-%m-%d %H:%M') if self.fecha else '',
            'nombres_completos': self.nombres_completos,
            'dni': self.dni,
            'profesion': self.profesion,
            'empresa': self.empresa,
            'cargo': self.cargo,
            'sector': self.sector,
            'fecha_nacimiento': self.fecha_nacimiento.strftime('%Y-%m-%d') if self.fecha_nacimiento else '',
            'estado': self.estado,
            'evaluador': self.evaluador or '',
            'observaciones': self.observaciones or '',
            'fecha_evaluacion': self.fecha_evaluacion.strftime('%Y-%m-%d %H:%M') if self.fecha_evaluacion else ''
        }

# Usuarios del sistema
USUARIOS = {
    'admin': {
        'password': 'admin123',
        'rol': 'administrador',
        'nombre': 'Administrador del Sistema'
    },
    'usuario': {
        'password': 'user123',
        'rol': 'usuario',
        'nombre': 'Usuario Regular'
    }
}

# Archivo para almacenar las solicitudes (simulando una base de datos)
SOLICITUDES_FILE = 'solicitudes.json'

def cargar_solicitudes():
    """Cargar solicitudes desde el archivo JSON"""
    if os.path.exists(SOLICITUDES_FILE):
        try:
            with open(SOLICITUDES_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (json.JSONDecodeError, FileNotFoundError):
            return []
    return []

def guardar_solicitudes(solicitudes):
    """Guardar solicitudes en el archivo JSON"""
    try:
        with open(SOLICITUDES_FILE, 'w', encoding='utf-8') as f:
            json.dump(solicitudes, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"Error al guardar solicitudes: {e}")
        return False

def obtener_siguiente_id():
    """Obtener el siguiente ID disponible"""
    solicitudes = cargar_solicitudes()
    if not solicitudes:
        return 1
    return max(s.get('id', 0) for s in solicitudes) + 1

def login_required(f):
    """Decorador para requerir login"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario' not in session:
            return jsonify({'error': 'No autorizado', 'redirect': '/login'}), 401
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    """Decorador para requerir permisos de administrador"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario' not in session:
            return jsonify({'error': 'No autorizado', 'redirect': '/login'}), 401
        if session.get('rol') != 'administrador':
            return jsonify({'error': 'Permisos insuficientes'}), 403
        return f(*args, **kwargs)
    return decorated_function

@app.route('/')
def index():
    """Ruta principal que renderiza el template de incorporación"""
    return render_template('incorporacion.html')

@app.route('/incorporacion')
def incorporacion():
    """Ruta específica para el proceso de incorporación"""
    return render_template('incorporacion.html')

@app.route('/api/login', methods=['POST'])
def login():
    """API endpoint para autenticación de usuarios"""
    try:
        data = request.get_json()
        usuario = data.get('usuario')
        password = data.get('password')
        
        if not usuario or not password:
            return jsonify({
                'success': False,
                'error': 'Usuario y contraseña son requeridos'
            }), 400
        
        # Verificar credenciales
        if usuario in USUARIOS and USUARIOS[usuario]['password'] == password:
            # Crear sesión
            session['usuario'] = usuario
            session['rol'] = USUARIOS[usuario]['rol']
            session['nombre'] = USUARIOS[usuario]['nombre']
            
            return jsonify({
                'success': True,
                'usuario': usuario,
                'rol': USUARIOS[usuario]['rol'],
                'nombre': USUARIOS[usuario]['nombre'],
                'message': 'Login exitoso'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Credenciales inválidas'
            }), 401
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/logout', methods=['POST'])
def logout():
    """API endpoint para cerrar sesión"""
    try:
        session.clear()
        return jsonify({
            'success': True,
            'message': 'Sesión cerrada correctamente'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/session', methods=['GET'])
def get_session():
    """API endpoint para obtener información de la sesión actual"""
    try:
        if 'usuario' in session:
            return jsonify({
                'authenticated': True,
                'usuario': session['usuario'],
                'rol': session['rol'],
                'nombre': session['nombre']
            })
        else:
            return jsonify({
                'authenticated': False
            })
    except Exception as e:
        return jsonify({
            'error': str(e)
        }), 500

@app.route('/api/solicitudes', methods=['GET'])
def obtener_solicitudes():
    """API endpoint para obtener todas las solicitudes"""
    try:
        solicitudes = Solicitud.query.all()
        return jsonify([solicitud.to_dict() for solicitud in solicitudes])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/solicitudes', methods=['POST'])
def crear_solicitud():
    """API endpoint para crear una nueva solicitud"""
    try:
        data = request.get_json()
        
        # Validar datos requeridos
        campos_requeridos = [
            'nombres_completos', 'dni', 'profesion', 'empresa', 
            'cargo', 'sector', 'fecha_nacimiento'
        ]
        
        for campo in campos_requeridos:
            if not data.get(campo):
                return jsonify({
                    'success': False, 
                    'error': f'El campo {campo} es requerido'
                }), 400
        
        # Verificar si ya existe una solicitud con el mismo DNI
        solicitud_existente = Solicitud.query.filter_by(dni=data.get('dni')).first()
        if solicitud_existente:
            return jsonify({
                'success': False,
                'error': 'Ya existe una solicitud con este DNI'
            }), 400
        
        # Convertir fecha de nacimiento
        try:
            fecha_nacimiento = datetime.strptime(data.get('fecha_nacimiento'), '%Y-%m-%d').date()
        except ValueError:
            return jsonify({
                'success': False,
                'error': 'Formato de fecha de nacimiento inválido'
            }), 400
        
        # Crear nueva solicitud
        nueva_solicitud = Solicitud(
            nombres_completos=data.get('nombres_completos'),
            dni=data.get('dni'),
            profesion=data.get('profesion'),
            empresa=data.get('empresa'),
            cargo=data.get('cargo'),
            sector=data.get('sector'),
            fecha_nacimiento=fecha_nacimiento
        )
        
        # Guardar en la base de datos
        db.session.add(nueva_solicitud)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'id': nueva_solicitud.id,
            'message': 'Solicitud creada correctamente'
        })
            
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/solicitudes/<int:solicitud_id>', methods=['PATCH'])
def actualizar_solicitud(solicitud_id):
    """API endpoint para actualizar una solicitud (evaluación)"""
    try:
        data = request.get_json()
        
        # Buscar la solicitud
        solicitud = Solicitud.query.get(solicitud_id)
        if not solicitud:
            return jsonify({
                'success': False,
                'error': 'Solicitud no encontrada'
            }), 404
        
        # Actualizar campos permitidos
        if 'estado' in data:
            solicitud.estado = data['estado']
        if 'evaluador' in data:
            solicitud.evaluador = data['evaluador']
        if 'observaciones' in data:
            solicitud.observaciones = data['observaciones']
        
        # Agregar fecha de evaluación
        solicitud.fecha_evaluacion = datetime.utcnow()
        
        # Guardar cambios
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Solicitud actualizada correctamente'
        })
            
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/solicitudes/<int:solicitud_id>', methods=['GET'])
def obtener_solicitud(solicitud_id):
    """API endpoint para obtener una solicitud específica"""
    try:
        solicitudes = cargar_solicitudes()
        
        for solicitud in solicitudes:
            if solicitud.get('id') == solicitud_id:
                return jsonify(solicitud)
        
        return jsonify({
            'error': 'Solicitud no encontrada'
        }), 404
        
    except Exception as e:
        return jsonify({
            'error': str(e)
        }), 500

@app.route('/api/solicitudes/<int:solicitud_id>', methods=['DELETE'])
def eliminar_solicitud(solicitud_id):
    """API endpoint para eliminar una solicitud"""
    try:
        solicitudes = cargar_solicitudes()
        
        # Filtrar la solicitud a eliminar
        solicitudes_filtradas = [s for s in solicitudes if s.get('id') != solicitud_id]
        
        if len(solicitudes_filtradas) == len(solicitudes):
            return jsonify({
                'success': False,
                'error': 'Solicitud no encontrada'
            }), 404
        
        if guardar_solicitudes(solicitudes_filtradas):
            return jsonify({
                'success': True,
                'message': 'Solicitud eliminada correctamente'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Error al eliminar la solicitud'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/descargar_solicitudes_excel')
def descargar_excel():
    """Endpoint para descargar las solicitudes en formato Excel"""
    try:
        solicitudes = Solicitud.query.all()
        
        if not solicitudes:
            return jsonify({
                'error': 'No hay solicitudes para descargar'
            }), 404
        
        # Convertir a diccionarios
        solicitudes_dict = [solicitud.to_dict() for solicitud in solicitudes]
        
        # Crear DataFrame
        df = pd.DataFrame(solicitudes_dict)
        
        # Reordenar columnas para mejor presentación (sin evaluador y observaciones)
        columnas_orden = [
            'id', 'fecha', 'nombres_completos', 'dni', 'profesion', 
            'empresa', 'cargo', 'sector', 'fecha_nacimiento', 'estado'
        ]
        
        # Filtrar solo las columnas que existen
        columnas_existentes = [col for col in columnas_orden if col in df.columns]
        df = df[columnas_existentes]
        
        # Renombrar columnas para mejor legibilidad
        nombres_columnas = {
            'id': 'ID',
            'fecha': 'Fecha Solicitud',
            'nombres_completos': 'Nombres Completos',
            'dni': 'DNI',
            'profesion': 'Profesión',
            'empresa': 'Empresa',
            'cargo': 'Cargo',
            'sector': 'Sector',
            'fecha_nacimiento': 'Fecha Nacimiento',
            'estado': 'Estado'
        }
        
        df = df.rename(columns=nombres_columnas)
        
        # Crear archivo Excel en memoria
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Solicitudes', index=False)
            
            # Obtener el workbook y worksheet para formatear
            workbook = writer.book
            worksheet = writer.sheets['Solicitudes']
            
            # Importar estilos de openpyxl para formateo
            from openpyxl.styles import Alignment, Font, PatternFill
            
            # Formatear encabezados
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Centrar la columna ID (columna A)
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet[f'A{row}']
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # Generar nombre de archivo con fecha
        fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_archivo = f'solicitudes_incorporacion_{fecha_actual}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nombre_archivo
        )
        
    except Exception as e:
        return jsonify({
            'error': f'Error al generar Excel: {str(e)}'
        }), 500

@app.route('/api/estadisticas')
def obtener_estadisticas():
    """API endpoint para obtener estadísticas de las solicitudes"""
    try:
        solicitudes = Solicitud.query.all()
        
        total = len(solicitudes)
        pendientes = len([s for s in solicitudes if s.estado == 'Pendiente'])
        aprobadas = len([s for s in solicitudes if s.estado == 'Aprobado'])
        rechazadas = len([s for s in solicitudes if s.estado == 'Rechazado'])
        
        # Estadísticas por sector
        sectores = {}
        for solicitud in solicitudes:
            sector = solicitud.sector or 'No especificado'
            sectores[sector] = sectores.get(sector, 0) + 1
        
        # Estadísticas por mes
        solicitudes_por_mes = {}
        for solicitud in solicitudes:
            if solicitud.fecha:
                mes_año = solicitud.fecha.strftime('%Y-%m')
                solicitudes_por_mes[mes_año] = solicitudes_por_mes.get(mes_año, 0) + 1
        
        return jsonify({
            'total': total,
            'pendientes': pendientes,
            'aprobadas': aprobadas,
            'rechazadas': rechazadas,
            'por_sector': sectores,
            'por_mes': solicitudes_por_mes
        })
        
    except Exception as e:
        return jsonify({
            'error': str(e)
        }), 500

@app.route('/api/status')
def api_status():
    """API endpoint para verificar el estado del servidor"""
    return jsonify({
        'status': 'active',
        'message': 'Servidor de incorporación funcionando correctamente',
        'timestamp': datetime.now().isoformat(),
        'version': '1.0.0'
    })

@app.errorhandler(404)
def not_found(error):
    """Manejador de errores 404"""
    return jsonify({
        'error': 'Endpoint no encontrado',
        'status': 404
    }), 404

@app.errorhandler(500)
def internal_error(error):
    """Manejador de errores 500"""
    return jsonify({
        'error': 'Error interno del servidor',
        'status': 500
    }), 500

# Inicialización para PythonAnywhere
try:
    os.makedirs('static/css', exist_ok=True)
    os.makedirs('static/images', exist_ok=True)
    os.makedirs('templates', exist_ok=True)
except:
    pass

# Crear las tablas de la base de datos
try:
    with app.app_context():
        db.create_all()
        print("✅ Base de datos SQLite inicializada correctamente")
except Exception as e:
    print(f"Error al inicializar BD: {e}")

if __name__ == '__main__':
    print("🚀 Iniciando servidor de incorporación...")
    print("📋 Sistema de gestión para nuevos asociados")
    print("🗄️ Base de datos: SQLite (solicitudes.db)")
    print("🌐 Accede a: http://localhost:5000")
    print("📊 API disponible en: http://localhost:5000/api/")
    
    # Ejecutar la aplicación
    app.run(debug=True, host='0.0.0.0', port=5000)